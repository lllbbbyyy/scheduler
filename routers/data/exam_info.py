from fastapi import APIRouter, Path, Depends, UploadFile
from tempfile import NamedTemporaryFile
from typing import Optional
from pydantic import BaseModel
from sqlalchemy.orm import Session
from .common_schema import Result, TimeStr
from .db_manage.models import Subject, Exam, ExamItem, LimitItem, Teacher, ScheduledItem
from .db_manage.database import get_db
from .scheduler import scheduler, scheduler_for_charge, scheduler_for_duty

import datetime
import openpyxl


class Item(BaseModel):
    exam_id: int
    exam_name: str
    exam_create_time: str


router = APIRouter()


@router.get("/exam_info", description='查询所有考试')
async def get_all_item(page: Optional[int] = None,
                       perPage: Optional[int] = None,
                       keywords: Optional[str] = '',
                       db: Session = Depends(get_db)):
    all_item = db.query(Exam).filter(Exam.name.like('%' + keywords + '%')).all()
    all_item.sort(key=lambda x:x.create_time)
    all_item.reverse()
    list_len = len(all_item)
    items_list = {'items': []}
    begin_index = (page - 1) * perPage
    for i in range(perPage):
        if begin_index + i >= list_len:
            break
        item = all_item[begin_index + i]
        items_list["items"].append(
            Item(exam_id=item.id,
                 exam_name=item.name,
                 exam_create_time=item.create_time.strftime(
                     "%Y-%m-%d %H:%M:%S")))
    items_list['total'] = list_len
    r = Result(data=items_list)
    return r


@router.put("/exam_info/file", description='通过文件增添考试')
async def add_item(file: UploadFile, db: Session = Depends(get_db)):
    contents = await file.read()
    with NamedTemporaryFile(mode='rb+', suffix='.xlsx') as tmp:
        tmp.write(contents)
        tmp.seek(0)
        wb = openpyxl.load_workbook(tmp)
    ws = wb.active
    exam = Exam(name=ws.title, create_time=datetime.datetime.now())
    if '高一' in ws.title:
        exam.grade = '高一'
    elif '高二' in ws.title:
        exam.grade = '高二'
    elif '高三' in ws.title:
        exam.grade = '高三'
    else:
        return Result(status=-1,
                      msg='未能查询到该考试的年级信息，请在该文件的sheet表1的title中包含考试年级高x')
    db.add(exam)
    db.commit()
    try:
        needed_add_exam_item = []
        for row in ws.iter_rows():
            exam_item = ExamItem(exam_id=exam.id)
            cnt = 0
            for cell in row:
                if cell.column == 1 and (not isinstance(cell.value, str)
                                         or cell.value[0] == '#'):
                    break
                if cell.column == 1 and isinstance(cell.value, str):
                    subject = db.query(Subject).filter(
                        Subject.name == cell.value).first()
                    if subject is None:
                        db.query(Exam).filter(Exam.id == exam.id).delete()
                        db.commit()
                        return Result(status=-1,
                                      msg="考试表中，科目：" + cell.value + "不存在")
                    exam_item.subject_id = subject.id
                    cnt += 1
                elif cell.column == 2:
                    exam_item.week = int(cell.value)
                    cnt += 1
                elif cell.column == 3:
                    exam_item.begin_time = str(TimeStr(str(cell.value)))
                    cnt += 1
                elif cell.column == 4:
                    exam_item.end_time = str(TimeStr(str(cell.value)))
                    cnt += 1
                elif cell.column == 5:
                    exam_item.needed_num = int(cell.value)
                    cnt += 1
            if cnt == 5:
                needed_add_exam_item.append(exam_item)
                duty_exam_item = exam_item.clone()
                duty_exam_item.type = 1
                duty_exam_item.needed_num = 1
                needed_add_exam_item.append(duty_exam_item)

        needed_add_limit_item = []
        if len(wb.sheetnames) >= 2:
            ws = wb[wb.sheetnames[1]]
            for row in ws.iter_rows():
                limit_item = LimitItem(exam_id=exam.id)
                cnt = 0
                for cell in row:
                    if cell.column == 1 and (not isinstance(cell.value, str)
                                             or cell.value[0] == '#'):
                        break
                    if cell.column == 1 and isinstance(cell.value, str):
                        teacher = db.query(Teacher).filter(
                            Teacher.name == cell.value).first()
                        if teacher is None:
                            db.query(Exam).filter(Exam.id == exam.id).delete()
                            return Result(status=-1,
                                          msg="限制条件表中，教师：" + cell.value +
                                          "不存在")
                        limit_item.teacher_id = teacher.id
                        cnt += 1
                    elif cell.column == 2:
                        limit_item.week = int(cell.value)
                        cnt += 1
                    elif cell.column == 3:
                        limit_item.begin_time = str(TimeStr(str(cell.value)))
                        cnt += 1
                    elif cell.column == 4:
                        limit_item.end_time = str(TimeStr(str(cell.value)))
                        cnt += 1
                if cnt == 4:
                    needed_add_limit_item.append(limit_item)
        db.add_all(needed_add_exam_item)
        db.add_all(needed_add_limit_item)
        db.commit()
        if not scheduler(exam):
            db.query(ExamItem).filter(ExamItem.exam_id == exam.id).delete()
            db.query(LimitItem).filter(LimitItem.exam_id == exam.id).delete()
            db.query(ScheduledItem).filter(
                ScheduledItem.exam_id == exam.id).delete()
            db.query(Exam).filter(Exam.id == exam.id).delete()
            db.commit()
            return Result(status=-1, msg='排考失败')
        if not scheduler_for_duty(exam):
            db.query(ExamItem).filter(ExamItem.exam_id == exam.id,
                                      ExamItem.type == 1).delete()
            db.query(ScheduledItem).filter(ScheduledItem.exam_id == exam.id,
                                           ScheduledItem.type == 1).delete()
            db.commit()
            return Result(status=-1, msg='监考排考成功，两处排考失败')
        scheduler_for_charge(exam)
    except Exception as e:
        db.query(ExamItem).filter(ExamItem.exam_id == exam.id).delete()
        db.query(LimitItem).filter(LimitItem.exam_id == exam.id).delete()
        db.query(ScheduledItem).filter(
            ScheduledItem.exam_id == exam.id).delete()
        db.query(Exam).filter(Exam.id == exam.id).delete()
        db.commit()
        raise e
    return Result()


@router.get("/exam_info/scheduled_items", description='查询已排考表条目信息')
async def get_all_item(exam_id: int, db: Session = Depends(get_db)):
    scheduled_items_list: list[ScheduledItem] = db.query(ScheduledItem).filter(
        ScheduledItem.exam_id == exam_id, ScheduledItem.type == 0).all()
    scheduled_items_dict = {}
    for scheduled_item in scheduled_items_list:
        time_segment = scheduled_item.begin_time + '-' + scheduled_item.end_time
        if time_segment not in scheduled_items_dict.keys():
            scheduled_items_dict[time_segment] = {}
        week = 'day' + str(scheduled_item.week)
        if week not in scheduled_items_dict[time_segment].keys():
            scheduled_items_dict[time_segment][week] = []
        scheduled_items_dict[time_segment][week].append({
            'name':
            db.query(Teacher).filter(
                Teacher.id == scheduled_item.teacher_id).first().name
        })
    sorted_items = sorted(scheduled_items_dict.items(), key=lambda x: x[0])
    r_data = {'items': []}
    day_set = set()
    #print(sorted_items)
    for sorted_item in sorted_items:
        begin_index = len(r_data['items'])
        is_first = True
        for week, name_list in sorted_item[1].items():
            day_set.add(week)
            i = begin_index
            for name_dict in name_list:
                if is_first:
                    d = {'time_segment': sorted_item[0]}
                    d[week] = name_dict['name']
                    r_data['items'].append(d)
                else:
                    r_data['items'][i][week] = name_dict['name']
                    i += 1
            is_first = False
    day_list = list(day_set)
    r_data['columns'] = []
    day_to_str = ['', '周一', '周二', '周三', '周四', '周五', '周六', '周日']
    for day in day_list:
        r_data['columns'].append({
            "name": day,
            "label": day_to_str[int(day[-1])],
            "type": "text",
            "quickEdit": True
        })
    r_data['columns'].sort(key=lambda x: x['name'])
    r_data['columns'].reverse()
    r_data['columns'].append({
        "name": "time_segment",
        "label": "时间段",
        "type": "text"
    })
    r_data['columns'].reverse()
    #print(r_data['items'])
    return r_data


@router.get("/exam_info/scheduled_duty_items", description='查询已排考两处表条目信息')
async def get_all_duty_item(exam_id: int, db: Session = Depends(get_db)):
    scheduled_items_list: list[ScheduledItem] = db.query(ScheduledItem).filter(
        ScheduledItem.exam_id == exam_id, ScheduledItem.type == 1).all()
    scheduled_items_dict = {}
    for scheduled_item in scheduled_items_list:
        time_segment = scheduled_item.begin_time + '-' + scheduled_item.end_time
        if time_segment not in scheduled_items_dict.keys():
            scheduled_items_dict[time_segment] = {}
        week = 'day' + str(scheduled_item.week)
        if week not in scheduled_items_dict[time_segment].keys():
            scheduled_items_dict[time_segment][week] = []
        scheduled_items_dict[time_segment][week].append({
            'name':
            db.query(Teacher).filter(
                Teacher.id == scheduled_item.teacher_id).first().name
        })
    sorted_items = sorted(scheduled_items_dict.items(), key=lambda x: x[0])
    r_data = {'items': []}
    day_set = set()
    #print(sorted_items)
    for sorted_item in sorted_items:
        begin_index = len(r_data['items'])
        is_first = True
        for week, name_list in sorted_item[1].items():
            day_set.add(week)
            i = begin_index
            for name_dict in name_list:
                if is_first:
                    d = {'time_segment': sorted_item[0]}
                    d[week] = name_dict['name']
                    r_data['items'].append(d)
                else:
                    r_data['items'][i][week] = name_dict['name']
                    i += 1
            is_first = False
    day_list = list(day_set)
    r_data['columns'] = []
    day_to_str = ['', '周一', '周二', '周三', '周四', '周五', '周六', '周日']
    for day in day_list:
        r_data['columns'].append({
            "name": day,
            "label": day_to_str[int(day[-1])],
            "type": "text",
            "quickEdit": True
        })
    r_data['columns'].sort(key=lambda x: x['name'])
    r_data['columns'].reverse()
    r_data['columns'].append({
        "name": "time_segment",
        "label": "时间段",
        "type": "text"
    })
    r_data['columns'].reverse()
    #print(r_data['items'])
    return r_data


@router.get("/exam_info/scheduled_charge_items", description='查询已排考行政表条目信息')
async def get_all_duty_item(exam_id: int, db: Session = Depends(get_db)):
    #这里0或者1都可以，1数据量小
    exam_item_list: list[ExamItem] = db.query(ExamItem).filter(
        ExamItem.exam_id == exam_id, ExamItem.type == 1).all()
    days_set = set()
    for exam_item in exam_item_list:
        days_set.add(exam_item.week)
    days_list: list[int] = list(days_set)
    days_list.sort()
    r_data = {'columns': [], 'items': [{}]}
    day_to_str = ['', '周一', '周二', '周三', '周四', '周五', '周六', '周日']
    for day in days_list:
        r_data['columns'].append({
            "name": 'day' + str(day),
            "label": day_to_str[day],
            "type": "text",
            "quickEdit": True
        })
    scheduler_item_list = db.query(ScheduledItem).filter(
        ScheduledItem.type == 2, ScheduledItem.exam_id == exam_id).all()
    for scheduler_item in scheduler_item_list:
        teacher = db.query(Teacher).filter(
            Teacher.id == scheduler_item.teacher_id).first()
        r_data['items'][0]['day' + str(scheduler_item.week)] = teacher.name
    return r_data


class DeleteItem(BaseModel):
    exam_id: int


@router.delete("/exam_info", description='单个删除数据')
async def item_delete(item: DeleteItem, db: Session = Depends(get_db)):
    db.query(ExamItem).filter(ExamItem.exam_id == item.exam_id).delete()
    db.query(LimitItem).filter(LimitItem.exam_id == item.exam_id).delete()
    db.query(ScheduledItem).filter(
        ScheduledItem.exam_id == item.exam_id).delete()
    db.query(Exam).filter(Exam.id == item.exam_id).delete()
    db.commit()
    return Result()


@router.delete("/exam_info/{exam_ids}", description='批量删除数据')
async def batch_delete(exam_ids: str = Path(..., title='通过,分隔'),
                       db: Session = Depends(get_db)):
    exam_id_list = exam_ids.split(',')
    for exam_id in exam_id_list:
        db.query(ExamItem).filter(ExamItem.exam_id == exam_id).delete()
        db.query(LimitItem).filter(LimitItem.exam_id == exam_id).delete()
        db.query(ScheduledItem).filter(
            ScheduledItem.exam_id == exam_id).delete()
        db.query(Exam).filter(Exam.id == exam_id).delete()
    db.commit()
    return Result()


@router.post("/exam_info/scheduled_items", description='修改排考表')
async def update_item(content: dict,
                      exam_id: int,
                      db: Session = Depends(get_db)):
    for i in range(len(content['rowsDiff'])):
        item = content['rowsDiff'][i]
        origin_item = content['rowsOrigin'][i]
        time_list = item['time_segment'].split('-')
        for k, v in item.items():
            if k == 'time_segment':
                continue
            #删除操作
            if v == '' and k in origin_item.keys() and origin_item[k] != '':
                teacher = db.query(Teacher).filter(
                    Teacher.name == origin_item[k]).first()
                db.query(ScheduledItem).filter(
                    ScheduledItem.exam_id == exam_id,
                    ScheduledItem.begin_time == time_list[0],
                    ScheduledItem.end_time == time_list[1],
                    ScheduledItem.teacher_id == teacher.id,
                    ScheduledItem.week == int(k[-1]),
                    ScheduledItem.type == 0).delete()
            #新增操作
            elif v != '' and (k not in origin_item.keys()
                              or origin_item[k] == ''):
                teacher = db.query(Teacher).filter(Teacher.name == v).first()
                exam = db.query(Exam).filter(Exam.id == exam_id).first()
                if teacher is None:
                    Result(status=-1, msg='教师：' + v + '不存在')
                db.add(
                    ScheduledItem(teacher_id=teacher.id,
                                  exam_id=exam_id,
                                  begin_time=time_list[0],
                                  end_time=time_list[1],
                                  week=int(k[-1]),
                                  grade=exam.grade))
            elif v != '' and k in origin_item.keys() and v != origin_item[k]:
                teacher_origin = db.query(Teacher).filter(
                    Teacher.name == origin_item[k]).first()
                teacher = db.query(Teacher).filter(Teacher.name == v).first()
                exam = db.query(Exam).filter(Exam.id == exam_id).first()
                if teacher is None:
                    return Result(status=-1, msg='教师：' + v + '不存在')
                db.query(ScheduledItem).filter(
                    ScheduledItem.exam_id == exam_id,
                    ScheduledItem.begin_time == time_list[0],
                    ScheduledItem.end_time == time_list[1],
                    ScheduledItem.teacher_id == teacher_origin.id,
                    ScheduledItem.week == int(k[-1]),
                    ScheduledItem.type == 0).delete()
                db.add(
                    ScheduledItem(teacher_id=teacher.id,
                                  exam_id=exam_id,
                                  begin_time=time_list[0],
                                  end_time=time_list[1],
                                  week=int(k[-1]),
                                  grade=exam.grade))
    db.commit()
    return Result()


@router.post("/exam_info/scheduled_duty_items", description='修改两处排考表')
async def update_item(content: dict,
                      exam_id: int,
                      db: Session = Depends(get_db)):
    for i in range(len(content['rowsDiff'])):
        item = content['rowsDiff'][i]
        origin_item = content['rowsOrigin'][i]
        time_list = item['time_segment'].split('-')
        for k, v in item.items():
            if k == 'time_segment':
                continue
            #删除操作
            if v == '' and k in origin_item.keys() and origin_item[k] != '':
                teacher = db.query(Teacher).filter(
                    Teacher.name == origin_item[k]).first()
                db.query(ScheduledItem).filter(
                    ScheduledItem.exam_id == exam_id,
                    ScheduledItem.begin_time == time_list[0],
                    ScheduledItem.end_time == time_list[1],
                    ScheduledItem.teacher_id == teacher.id,
                    ScheduledItem.week == int(k[-1]),
                    ScheduledItem.type == 1).delete()
            #新增操作
            elif v != '' and (k not in origin_item.keys()
                              or origin_item[k] == ''):
                teacher = db.query(Teacher).filter(Teacher.name == v).first()
                exam = db.query(Exam).filter(Exam.id == exam_id).first()
                if teacher is None:
                    Result(status=-1, msg='教师：' + v + '不存在')
                db.add(
                    ScheduledItem(teacher_id=teacher.id,
                                  exam_id=exam_id,
                                  begin_time=time_list[0],
                                  end_time=time_list[1],
                                  week=int(k[-1]),
                                  grade=exam.grade,
                                  type=1))
            elif v != '' and k in origin_item.keys() and v != origin_item[k]:
                teacher_origin = db.query(Teacher).filter(
                    Teacher.name == origin_item[k]).first()
                teacher = db.query(Teacher).filter(Teacher.name == v).first()
                exam = db.query(Exam).filter(Exam.id == exam_id).first()
                if teacher is None:
                    return Result(status=-1, msg='教师：' + v + '不存在')
                db.query(ScheduledItem).filter(
                    ScheduledItem.exam_id == exam_id,
                    ScheduledItem.begin_time == time_list[0],
                    ScheduledItem.end_time == time_list[1],
                    ScheduledItem.teacher_id == teacher_origin.id,
                    ScheduledItem.week == int(k[-1]),
                    ScheduledItem.type == 1).delete()
                db.add(
                    ScheduledItem(teacher_id=teacher.id,
                                  exam_id=exam_id,
                                  begin_time=time_list[0],
                                  end_time=time_list[1],
                                  week=int(k[-1]),
                                  grade=exam.grade,
                                  type=1))
    db.commit()
    return Result()


@router.post("/exam_info/scheduled_charge_items", description='修改行政表')
async def update_item(content: dict,
                      exam_id: int,
                      db: Session = Depends(get_db)):
    for i in range(len(content['rowsDiff'])):
        item = content['rowsDiff'][i]
        origin_item = content['rowsOrigin'][i]
        for k, v in item.items():
            #删除操作
            if v == '' and k in origin_item.keys() and origin_item[k] != '':
                teacher = db.query(Teacher).filter(
                    Teacher.name == origin_item[k]).first()
                db.query(ScheduledItem).filter(
                    ScheduledItem.exam_id == exam_id,
                    ScheduledItem.teacher_id == teacher.id,
                    ScheduledItem.week == int(k[-1]),
                    ScheduledItem.type == 2).delete()
            #新增操作
            elif v != '' and (k not in origin_item.keys()
                              or origin_item[k] == ''):
                teacher = db.query(Teacher).filter(Teacher.name == v).first()
                exam = db.query(Exam).filter(Exam.id == exam_id).first()
                if teacher is None:
                    Result(status=-1, msg='教师：' + v + '不存在')
                db.add(
                    ScheduledItem(teacher_id=teacher.id,
                                  exam_id=exam_id,
                                  week=int(k[-1]),
                                  begin_time='00:00',
                                  end_time='00:01',
                                  grade=exam.grade,
                                  type=2))
            elif v != '' and k in origin_item.keys() and v != origin_item[k]:
                teacher_origin = db.query(Teacher).filter(
                    Teacher.name == origin_item[k]).first()
                teacher = db.query(Teacher).filter(Teacher.name == v).first()
                exam = db.query(Exam).filter(Exam.id == exam_id).first()
                if teacher is None:
                    return Result(status=-1, msg='教师：' + v + '不存在')
                db.query(ScheduledItem).filter(
                    ScheduledItem.exam_id == exam_id,
                    ScheduledItem.teacher_id == teacher_origin.id,
                    ScheduledItem.week == int(k[-1]),
                    ScheduledItem.type == 2).delete()
                db.add(
                    ScheduledItem(teacher_id=teacher.id,
                                  exam_id=exam_id,
                                  begin_time='00:00',
                                  end_time='00:01',
                                  week=int(k[-1]),
                                  grade=exam.grade,
                                  type=2))
    db.commit()
    return Result()


# @router.put("/subject_info", description='增添科目信息')
# async def add_item(item: Item, db: Session = Depends(get_db)):
#     db_item = Subject(name=item.subject_name)
#     if db.query(Subject).filter(Subject.name == db_item.name).first():
#         return Result(status=-1, msg='已有该科目')
#     db.add(db_item)
#     db.commit()
#     db.refresh(db_item)
#     return Result()